"""
Report generator: PDF and Excel exports for marketing teams.
Uses reportlab for PDF and openpyxl for Excel.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

# Output directory under project data
REPORTS_DIR = Path(__file__).resolve().parent.parent / "data" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


class ReportGenerator:
    """Generate PDF and Excel reports from sentiment analytics data."""

    def __init__(self, output_dir: Path = None):
        self.output_dir = output_dir or REPORTS_DIR

    def _ensure_reports_dir(self) -> Path:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        return self.output_dir

    def _get_report_data(self, analyzer) -> Dict:
        """Get full report dict from SentimentAnalyzer."""
        return {
            "timestamp": datetime.now().isoformat(),
            "total_records": len(analyzer.df),
            "overall_sentiment": analyzer.get_overall_sentiment_score(),
            "feature_sentiments": analyzer.get_feature_sentiments(),
            "model_sentiments": analyzer.get_model_sentiments(),
            "top_features": analyzer.get_top_features(n=20),
        }

    def generate_pdf(
        self,
        report: Dict,
        filename: Optional[str] = None,
    ) -> Path:
        """Generate PDF report. report = dict with overall_sentiment, feature_sentiments, model_sentiments, top_features."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak,
            )
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install: pip install reportlab")

        self._ensure_reports_dir()
        fname = filename or f"sentiment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = self.output_dir / fname

        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=inch,
            leftMargin=inch,
            topMargin=inch,
            bottomMargin=inch,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=12,
        )
        h2_style = ParagraphStyle(
            "CustomH2",
            parent=styles["Heading2"],
            fontSize=14,
            spaceAfter=8,
        )
        story = []

        story.append(Paragraph("Redmi Sentiment Analysis – Marketing Report", title_style))
        story.append(Paragraph(f"Generated: {report.get('timestamp', 'N/A')}", styles["Normal"]))
        story.append(Spacer(1, 0.3 * inch))

        overall = report.get("overall_sentiment") or {}
        story.append(Paragraph("Overall Sentiment", h2_style))
        overall_data = [
            ["Metric", "Value"],
            ["Sentiment Score", f"{overall.get('sentiment_score', 0):.3f}"],
            ["Positive", str(overall.get("positive", 0))],
            ["Negative", str(overall.get("negative", 0))],
            ["Neutral", str(overall.get("neutral", 0))],
            ["Total", str(overall.get("total", 0))],
        ]
        t = Table(overall_data, colWidths=[2.5 * inch, 2 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.25 * inch))

        features = report.get("feature_sentiments") or {}
        if features:
            story.append(Paragraph("Feature Sentiment Scores", h2_style))
            rows = [["Feature", "Score", "Positive", "Negative", "Total"]]
            for feat, data in sorted(features.items(), key=lambda x: x[1].get("sentiment_score", 0), reverse=True)[:15]:
                rows.append([
                    feat,
                    f"{data.get('sentiment_score', 0):.3f}",
                    str(data.get("positive", 0)),
                    str(data.get("negative", 0)),
                    str(data.get("total", 0)),
                ])
            t2 = Table(rows, colWidths=[1.4 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 0.7 * inch])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2ecc71")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(t2)
            story.append(Spacer(1, 0.25 * inch))

        models = report.get("model_sentiments") or {}
        if models:
            story.append(Paragraph("Sentiment by Model", h2_style))
            rows = [["Model", "Score", "Positive", "Negative", "Total"]]
            for model, data in list(models.items())[:10]:
                rows.append([
                    model[:30],
                    f"{data.get('sentiment_score', 0):.3f}",
                    str(data.get("positive", 0)),
                    str(data.get("negative", 0)),
                    str(data.get("total", 0)),
                ])
            t3 = Table(rows, colWidths=[2 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 0.7 * inch])
            t3.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9b59b6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(t3)

        doc.build(story)
        logger.info("PDF report saved: %s", filepath)
        return filepath

    def generate_excel(
        self,
        report: Dict,
        filename: Optional[str] = None,
    ) -> Path:
        """Generate Excel report. report = same structure as for PDF."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install: pip install openpyxl")

        self._ensure_reports_dir()
        fname = filename or f"sentiment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / fname

        wb = Workbook()
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["Redmi Sentiment Analysis – Marketing Report"])
        ws1.append(["Generated", report.get("timestamp", "")])
        ws1.append([])
        overall = report.get("overall_sentiment") or {}
        ws1.append(["Overall Sentiment", ""])
        ws1.append(["Sentiment Score", overall.get("sentiment_score", 0)])
        ws1.append(["Positive", overall.get("positive", 0)])
        ws1.append(["Negative", overall.get("negative", 0)])
        ws1.append(["Neutral", overall.get("neutral", 0)])
        ws1.append(["Total", overall.get("total", 0)])
        ws1["A1"].font = Font(bold=True, size=14)

        # Sheet 2: Feature Sentiments
        ws2 = wb.create_sheet("Feature Sentiments")
        ws2.append(["Feature", "Sentiment Score", "Positive", "Negative", "Neutral", "Total"])
        for r in range(1, 7):
            ws2.cell(1, r).fill = header_fill
            ws2.cell(1, r).font = header_font
        for feat, data in sorted(
            (report.get("feature_sentiments") or {}).items(),
            key=lambda x: x[1].get("sentiment_score", 0),
            reverse=True,
        ):
            ws2.append([
                feat,
                data.get("sentiment_score", 0),
                data.get("positive", 0),
                data.get("negative", 0),
                data.get("neutral", 0),
                data.get("total", 0),
            ])

        # Sheet 3: Model Sentiments
        ws3 = wb.create_sheet("Model Sentiments")
        ws3.append(["Model", "Sentiment Score", "Positive", "Negative", "Neutral", "Total"])
        for r in range(1, 7):
            ws3.cell(1, r).fill = header_fill
            ws3.cell(1, r).font = header_font
        for model, data in (report.get("model_sentiments") or {}).items():
            ws3.append([
                model,
                data.get("sentiment_score", 0),
                data.get("positive", 0),
                data.get("negative", 0),
                data.get("neutral", 0),
                data.get("total", 0),
            ])

        # Sheet 4: Top Features
        ws4 = wb.create_sheet("Top Features")
        ws4.append(["Feature", "Mentions"])
        ws4.cell(1, 1).fill = header_fill
        ws4.cell(1, 1).font = header_font
        ws4.cell(1, 2).fill = header_fill
        ws4.cell(1, 2).font = header_font
        for feat, count in (report.get("top_features") or {}).items():
            ws4.append([feat, count])

        wb.save(str(filepath))
        logger.info("Excel report saved: %s", filepath)
        return filepath

    def generate_all(self, analyzer, pdf_name: Optional[str] = None, excel_name: Optional[str] = None) -> Dict[str, Path]:
        """Generate both PDF and Excel from SentimentAnalyzer. Returns dict with 'pdf' and 'excel' paths."""
        report = self._get_report_data(analyzer)
        return {
            "pdf": self.generate_pdf(report, filename=pdf_name),
            "excel": self.generate_excel(report, filename=excel_name),
        }
